# -*- coding: utf-8 -*-
"""
Отчёты: Excel (.xlsx) и PDF — «Карточка учёта рабочего времени и выплат».
Кириллица в PDF работает через встроенный шрифт DejaVuSans.ttf из assets/.
"""
import os
import datetime as dt

from timecard_core import (Totals, WEEKDAYS_RU_SHORT, fmt_time, fmt_hm_short,
                           week_start, month_title)

HERE = os.path.dirname(os.path.abspath(__file__))
FONT_CANDIDATES = [
    os.path.join(HERE, "assets", "DejaVuSans.ttf"),
    os.path.join(HERE, "..", "windows", "assets", "DejaVuSans.ttf"),
    os.path.join(HERE, "..", "assets", "DejaVuSans.ttf"),
    r"C:\Windows\Fonts\arial.ttf",
    r"C:\Windows\Fonts\segoeui.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
]
FONT_BOLD_CANDIDATES = [
    os.path.join(HERE, "assets", "DejaVuSans-Bold.ttf"),
    os.path.join(HERE, "..", "windows", "assets", "DejaVuSans-Bold.ttf"),
    r"C:\Windows\Fonts\arialbd.ttf",
    r"C:\Windows\Fonts\segoeuib.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
]

HEADERS = ["Дата", "День", "Время работы", "Кол-во часов", "Доп. время", "Доп. часы",
           "Объём и качество произведённых работ", "Оплата за день", "Доп. работы",
           "Премия", "Итого к выплате"]


def _row(e):
    d = e.date_obj
    time_s = ("с %s по %s" % (fmt_time(e.start), fmt_time(e.end))
              if e.start is not None and e.end is not None else "")
    if e.lunch_on and e.lunch_min:
        time_s += " (обед %d мин)" % e.lunch_min
    xt = ("с %s по %s" % (fmt_time(e.extra_start), fmt_time(e.extra_end))
          if e.extra_on and e.extra_start is not None and e.extra_end is not None else "")
    works = e.works.replace("\n", "; ")
    if e.extra_on and e.extra_works:
        works = (works + " || доп.: " + e.extra_works.replace("\n", "; ")).strip(" |")
    return [
        "%02d.%02d.%d" % (d.day, d.month, d.year),
        WEEKDAYS_RU_SHORT[d.weekday()],
        time_s,
        fmt_hm_short(e.work_min) if e.work_min else "",
        xt,
        fmt_hm_short(e.extra_min) if e.extra_min else "",
        works,
        round(e.day_pay) if e.day_pay else "",
        round(e.extra_pay) if e.extra_pay else "",
        round(e.bonus) if e.bonus else "",
        round(e.total_pay) if e.total_pay else "",
    ]


# ---------------------------------------------------------------------- EXCEL
def export_xlsx(path, title, days, meta=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Для экспорта в Excel нужна библиотека openpyxl:\n"
                          "pip install openpyxl")
    meta = meta or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Табель"

    accent = PatternFill("solid", fgColor="5B9BD5")
    light = PatternFill("solid", fgColor="DEE7F0")
    week_fill = PatternFill("solid", fgColor="E3EDF7")
    total_fill = PatternFill("solid", fgColor="C9DFF2")
    thin = Side(style="thin", color="C3D2E0")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncol = len(HEADERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14, color="22313F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    sub = []
    if meta.get("employee"):
        sub.append("Работник: " + meta["employee"])
    if meta.get("organization"):
        sub.append(meta["organization"])
    sub.append("Ставка: %s ₽/час" % meta.get("rate", "250"))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    c = ws.cell(row=2, column=2 - 1, value="   •   ".join(sub))
    c.alignment = Alignment(horizontal="center")
    c.font = Font(size=10, color="6B7C8C")

    hr = 4
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = accent
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = box
    ws.row_dimensions[hr].height = 34

    r = hr + 1
    wk_buf = []

    def flush_week():
        nonlocal r
        t = Totals(wk_buf)
        if not t.worked_days:
            return
        vals = ["", "", "Итого за неделю", fmt_hm_short(t.work_min), "",
                fmt_hm_short(t.extra_min), "отработано дней: %d" % t.worked_days,
                round(t.day_pay), round(t.extra_pay), round(t.bonus), round(t.total_pay)]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.fill = week_fill
            c.font = Font(bold=True, size=10)
            c.border = box
        r += 1

    for e in days:
        if not e.is_empty:
            for i, v in enumerate(_row(e), 1):
                c = ws.cell(row=r, column=i, value=v)
                c.border = box
                c.alignment = Alignment(
                    horizontal="left" if i == 7 else "center",
                    vertical="center", wrap_text=(i == 7))
                if i >= 8 and isinstance(v, (int, float)):
                    c.number_format = '# ##0 \u20bd'
            if e.date_obj.weekday() >= 5:
                for i in range(1, ncol + 1):
                    ws.cell(row=r, column=i).fill = light
            r += 1
        wk_buf.append(e)
        if e.date_obj.weekday() == 6:
            flush_week()
            wk_buf = []
    if wk_buf:
        flush_week()

    t = Totals(days)
    vals = ["", "", "ВСЕГО", fmt_hm_short(t.work_min), "", fmt_hm_short(t.extra_min),
            "отработано дней: %d, всего %s" % (t.worked_days, fmt_hm_short(t.total_min)),
            round(t.day_pay), round(t.extra_pay), round(t.bonus), round(t.total_pay)]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.fill = total_fill
        c.font = Font(bold=True, size=11)
        c.border = box
        if i >= 8 and isinstance(v, (int, float)):
            c.number_format = '# ##0 \u20bd'
    ws.row_dimensions[r].height = 24

    widths = [12, 7, 20, 13, 18, 11, 52, 15, 14, 11, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A%d" % (hr + 1)
    ws.page_setup.orientation = "landscape"
    wb.save(path)
    return path


# ------------------------------------------------------------------------ PDF
def _register_font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    reg = pdfmetrics.getRegisteredFontNames()
    if "Tab" in reg:
        return "Tab", ("TabB" if "TabB" in reg else "Tab")
    regular = next((p for p in FONT_CANDIDATES if os.path.exists(p)), None)
    bold = next((p for p in FONT_BOLD_CANDIDATES if os.path.exists(p)), None)
    if not regular:
        raise ImportError("Не найден шрифт с кириллицей (DejaVuSans.ttf). "
                          "Положите его в папку assets рядом с программой.")
    pdfmetrics.registerFont(TTFont("Tab", regular))
    if bold:
        pdfmetrics.registerFont(TTFont("TabB", bold))
        return "Tab", "TabB"
    return "Tab", "Tab"


def export_pdf(path, title, days, meta=None):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
        from reportlab.lib.styles import ParagraphStyle
    except ImportError:
        raise ImportError("Для экспорта в PDF нужна библиотека reportlab:\n"
                          "pip install reportlab")
    meta = meta or {}
    fn, fb = _register_font()

    st_title = ParagraphStyle("t", fontName=fb, fontSize=14, leading=18,
                              textColor=colors.HexColor("#22313F"), alignment=1)
    st_sub = ParagraphStyle("s", fontName=fn, fontSize=9, leading=12,
                            textColor=colors.HexColor("#6B7C8C"), alignment=1)
    st_cell = ParagraphStyle("c", fontName=fn, fontSize=7.6, leading=9.4)
    st_head = ParagraphStyle("h", fontName=fb, fontSize=7.8, leading=9.6,
                             textColor=colors.white, alignment=1)

    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm,
                            title=title)
    story = [Paragraph(title, st_title)]
    sub = []
    if meta.get("employee"):
        sub.append("Работник: " + meta["employee"])
    if meta.get("organization"):
        sub.append(meta["organization"])
    sub.append("Ставка: %s ₽/час" % meta.get("rate", "250"))
    story += [Spacer(1, 3 * mm), Paragraph("   •   ".join(sub), st_sub),
              Spacer(1, 4 * mm)]

    data = [[Paragraph(h, st_head) for h in HEADERS]]
    styles = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#5B9BD5")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C3D2E0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 1), (5, -1), "CENTER"),
        ("ALIGN", (7, 1), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, 1), (-1, -1), fn),
        ("FONTSIZE", (0, 1), (-1, -1), 7.6),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]

    ri = 1
    wk_buf = []

    def flush_week():
        nonlocal ri
        t = Totals(wk_buf)
        if not t.worked_days:
            return
        data.append(["", "", "Итого за неделю", fmt_hm_short(t.work_min), "",
                     fmt_hm_short(t.extra_min), "отработано дней: %d" % t.worked_days,
                     "%d" % round(t.day_pay), "%d" % round(t.extra_pay),
                     "%d" % round(t.bonus), "%d" % round(t.total_pay)])
        styles.append(("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#E3EDF7")))
        styles.append(("FONTNAME", (0, ri), (-1, ri), fb))
        ri += 1

    for e in days:
        if not e.is_empty:
            r = _row(e)
            r[6] = Paragraph(str(r[6]), st_cell)
            r = [("%d" % v if isinstance(v, (int, float)) else v) for v in r]
            data.append(r)
            if e.date_obj.weekday() >= 5:
                styles.append(("BACKGROUND", (0, ri), (-1, ri),
                               colors.HexColor("#EDF3F9")))
            ri += 1
        wk_buf.append(e)
        if e.date_obj.weekday() == 6:
            flush_week()
            wk_buf = []
    if wk_buf:
        flush_week()

    t = Totals(days)
    data.append(["", "", "ВСЕГО", fmt_hm_short(t.work_min), "", fmt_hm_short(t.extra_min),
                 "отработано дней: %d, всего %s" % (t.worked_days,
                                                    fmt_hm_short(t.total_min)),
                 "%d" % round(t.day_pay), "%d" % round(t.extra_pay),
                 "%d" % round(t.bonus), "%d" % round(t.total_pay)])
    styles += [("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#C9DFF2")),
               ("FONTNAME", (0, ri), (-1, ri), fb),
               ("FONTSIZE", (0, ri), (-1, ri), 9)]

    widths = [17, 10, 30, 17, 26, 15, 88, 22, 21, 16, 24]
    total_w = sum(widths)
    avail = landscape(A4)[0] - 20 * mm
    col_w = [w / total_w * avail for w in widths]

    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle(styles))
    story.append(tbl)
    story += [Spacer(1, 6 * mm),
              Paragraph("Сформировано программой «Табель» — %s" %
                        dt.datetime.now().strftime("%d.%m.%Y %H:%M"), st_sub)]
    doc.build(story)
    return path


# ------------------------------------------------------------------ текст/CSV
def export_text(days, title, meta=None):
    """Простой текстовый отчёт — для кнопки «Поделиться» на Android."""
    t = Totals(days)
    lines = [title, "-" * 44]
    for e in days:
        if e.is_empty:
            continue
        d = e.date_obj
        s = "%s %02d.%02d  %s" % (WEEKDAYS_RU_SHORT[d.weekday()], d.day, d.month,
                                  fmt_hm_short(e.work_min))
        if e.extra_min:
            s += " +доп %s" % fmt_hm_short(e.extra_min)
        s += " = %d р." % round(e.total_pay)
        lines.append(s)
        if e.works:
            lines.append("    " + e.works.replace("\n", "; "))
    lines += ["-" * 44,
              "Дней: %d   Часы: %s   Доп.: %s" % (t.worked_days,
                                                  fmt_hm_short(t.work_min),
                                                  fmt_hm_short(t.extra_min)),
              "Оплата: %d + доп %d + премия %d" % (round(t.day_pay), round(t.extra_pay),
                                                   round(t.bonus)),
              "ИТОГО: %d руб." % round(t.total_pay)]
    return "\n".join(lines)
